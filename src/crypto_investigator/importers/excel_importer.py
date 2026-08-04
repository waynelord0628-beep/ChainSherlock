from pathlib import Path
from typing import Mapping

from openpyxl import load_workbook
import pandas as pd

from crypto_investigator.importers.base import ImportBatch
from crypto_investigator.importers.mapping import MappingEngine


class ExcelImporter:
    def __init__(self, mapping_engine: MappingEngine | None = None) -> None:
        self.mapping_engine = mapping_engine or MappingEngine()

    def load(
        self,
        path: Path,
        column_overrides: Mapping[str, str] | None = None,
    ) -> ImportBatch:
        if path.suffix.casefold() == ".xlsx":
            frame = self._load_xlsx(path)
        else:
            frame = pd.read_excel(path, dtype=object, engine="xlrd", keep_default_na=False)
        mapping = self.mapping_engine.resolve(frame.columns, column_overrides)
        records = tuple(
            mapping.apply(self._clean_row(row)) for row in frame.to_dict(orient="records")
        )
        return ImportBatch(path, records, mapping)

    @staticmethod
    def _load_xlsx(path: Path) -> pd.DataFrame:
        workbook = load_workbook(path, read_only=True, data_only=False)
        try:
            worksheet = workbook.active
            rows = list(worksheet.iter_rows(values_only=True))
        finally:
            workbook.close()
        if not rows:
            return pd.DataFrame()
        columns = [str(value) if value is not None else "" for value in rows[0]]
        return pd.DataFrame(rows[1:], columns=columns)

    @staticmethod
    def _clean_row(row: dict[str, object]) -> dict[str, object]:
        return {
            key: None if pd.isna(value) or value == "" else value
            for key, value in row.items()
        }

